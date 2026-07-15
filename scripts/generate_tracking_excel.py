#!/usr/bin/env python3
"""Generate wealth-app tracking spec Excel from JSON."""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing dependency: openpyxl. Install with: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

COLUMNS = [
    "埋点分类",
    "事件名称",
    "*事件英文名称\nevent_name",
    "页面名称",
    "*页面英文名称\npage_name",
    "模块名称",
    "*模块英文名称\nmodule_name",
    "元素名称",
    "*元素英文名称\nelement_name",
    "*元素位置\nlocation",
    "*产品代码\nproduct_code",
    "*备注\nremarks",
    "解释说明",
    "老埋点关联\neventId",
    "新增或修改日期",
]

HEADER_FILL = PatternFill("solid", fgColor="ED7D31")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGN = Alignment(vertical="top", wrap_text=True)

COL_WIDTHS = [10, 26, 36, 26, 36, 14, 20, 20, 28, 12, 14, 22, 24, 18, 16]

CATEGORY = {"view": "浏览事件", "click": "点击事件", "show": "曝光事件"}

_SCRIPTS = Path(__file__).resolve().parent
if str(_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS))
from event_name_cn import check_spec, normalize_event_cn, normalize_events_dict  # noqa: E402

try:
    from lookup_event_names import is_company_prefix  # noqa: E402
except ImportError:  # pragma: no cover
    is_company_prefix = None  # type: ignore


def warn_unknown_event_prefixes(document: dict) -> None:
    """Warn if event_name_en prefix is outside company event-name dictionary."""
    if is_company_prefix is None:
        return

    def check_en(name_en: str, label: str) -> None:
        if not name_en:
            return
        en = name_en.split("\n")[0].strip()
        m = re.match(r"(.+)_(view|click|show)$", en, re.I)
        if not m:
            print(f"[event_prefix WARN] {label}: 无法解析前缀 `{en}`", file=sys.stderr)
            return
        prefix = m.group(1)
        if not is_company_prefix(prefix):
            print(
                f"[event_prefix WARN] {label}: `{en}` 前缀 `{prefix}` 不在公司事件名称字典。"
                f" 确认单默认须用事件名称字典中的前缀（如 wealth_home / wealth_fund）；"
                f" 仅当用户明示「用户指定（非字典）」时可忽略。",
                file=sys.stderr,
            )

    pages = document.get("pages") or [document]
    for page_spec in pages:
        if "page" not in page_spec and "events" not in page_spec:
            continue
        page = page_spec.get("page") or {}
        label = page.get("name_en") or page.get("name_cn") or "?"
        for et, cfg in (page_spec.get("events") or {}).items():
            if isinstance(cfg, dict):
                check_en(cfg.get("name_en", ""), f"{label}/{et}")
        for key in ("view_events", "click_events", "show_events"):
            for item in page_spec.get(key) or []:
                check_en(item.get("event_name_en", ""), f"{label}/{item.get('interaction_id', key)}")


def load_spec(path: Path) -> dict:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def get_events_config(spec: dict) -> dict:
    page = spec["page"]
    events = spec.get("events", {})
    view_cn = events.get("view", {}).get("name_cn") or page.get(
        "view_event_cn", f"{page.get('product_line', '')}浏览事件".strip()
    )
    click_cn = events.get("click", {}).get("name_cn") or page.get(
        "click_event_cn", f"{page.get('product_line', '')}点击事件".strip()
    )
    show_cn = events.get("show", {}).get("name_cn") or page.get(
        "show_event_cn", f"{page.get('product_line', '')}曝光事件".strip()
    )
    return {
        "view": {
            "name_cn": normalize_event_cn(view_cn, "view"),
            "name_en": events.get("view", {}).get("name_en")
            or page.get("view_event_en", "wealth_home_view"),
        },
        "click": {
            "name_cn": normalize_event_cn(click_cn, "click"),
            "name_en": events.get("click", {}).get("name_en")
            or page.get("click_event_en", "wealth_home_click"),
        },
        "show": {
            "name_cn": normalize_event_cn(show_cn, "show"),
            "name_en": events.get("show", {}).get("name_en")
            or page.get("show_event_en", "wealth_home_show"),
        },
    }


def build_rows(spec: dict) -> list[list]:
    page = spec["page"]
    page_cn = page["name_cn"]
    page_en = page["name_en"]
    product_code = page.get("product_code", "")
    page_remarks = page.get("remarks", "")
    default_date = page.get("change_date", date.today().isoformat())
    events_cfg = get_events_config(spec)

    rows: list[list] = []

    def remarks(*parts: str) -> str:
        return " | ".join(p for p in parts if p)

    for item in spec.get("view_events", []):
        rows.append([
            CATEGORY["view"],
            item.get("event_name_cn") or events_cfg["view"]["name_cn"],
            item.get("event_name_en") or events_cfg["view"]["name_en"],
            page_cn,
            page_en,
            "", "", "", "",
            item.get("location", ""),
            product_code,
            remarks(page_remarks, item.get("remarks", "")),
            item.get("explanation", ""),
            item.get("legacy_event_id", ""),
            item.get("change_date", default_date),
        ])

    for item in spec.get("click_events", []):
        rows.append([
            CATEGORY["click"],
            item.get("event_name_cn") or events_cfg["click"]["name_cn"],
            item.get("event_name_en") or events_cfg["click"]["name_en"],
            page_cn,
            page_en,
            item.get("module_cn", ""),
            item.get("module_en", ""),
            item.get("element_cn", ""),
            item.get("element_en", ""),
            item.get("location", ""),
            product_code,
            remarks(page_remarks, item.get("remarks", "")),
            item.get("explanation", ""),
            item.get("legacy_event_id", ""),
            item.get("change_date", default_date),
        ])

    for item in spec.get("show_events", []):
        rows.append([
            CATEGORY["show"],
            item.get("event_name_cn") or events_cfg["show"]["name_cn"],
            item.get("event_name_en") or events_cfg["show"]["name_en"],
            page_cn,
            page_en,
            item.get("module_cn", ""),
            item.get("module_en", ""),
            item.get("element_cn", ""),
            item.get("element_en", ""),
            item.get("location", ""),
            product_code,
            remarks(page_remarks, item.get("remarks", "")),
            item.get("explanation", ""),
            item.get("legacy_event_id", ""),
            item.get("change_date", default_date),
        ])

    return rows


def merge_event_columns(ws, rows: list[list]) -> None:
    merge_cols = (2, 3)
    n = len(rows)
    if n == 0:
        return
    start = 2
    for i in range(1, n):
        curr_row = 2 + i
        prev_vals = tuple(rows[i - 1][c - 1] for c in merge_cols)
        curr_vals = tuple(rows[i][c - 1] for c in merge_cols)
        if curr_vals != prev_vals:
            end = 2 + i - 1
            for col in merge_cols:
                if end > start:
                    ws.merge_cells(start_row=start, start_column=col, end_row=end, end_column=col)
            start = curr_row
    end = 2 + n - 1
    for col in merge_cols:
        if end > start:
            ws.merge_cells(start_row=start, start_column=col, end_row=end, end_column=col)


def write_excel(rows: list[list], output: Path, sheet_name: str = "埋点提报") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    ws.append(COLUMNS)
    ws.row_dimensions[1].height = 36
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    for row in rows:
        ws.append(row)

    merge_event_columns(ws, rows)

    for idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(COLUMNS)):
        for cell in row:
            cell.alignment = BODY_ALIGN

    ws.freeze_panes = "A2"
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def collect_rows_from_document(document: dict) -> tuple[list[list], str]:
    if "pages" in document:
        sheet_name = document.get("sheet_name", "埋点提报")
        default_date = document.get("change_date", date.today().isoformat())
        product_line = document.get("product_line", "")
        all_rows: list[list] = []
        for page_spec in document["pages"]:
            page = page_spec["page"]
            spec = {
                "page": {
                    **page,
                    "sheet_name": sheet_name,
                    "product_line": product_line,
                    "change_date": page.get("change_date", default_date),
                },
                "events": page_spec.get("events", {}),
                "view_events": page_spec.get("view_events", []),
                "click_events": page_spec.get("click_events", []),
                "show_events": page_spec.get("show_events", []),
            }
            all_rows.extend(build_rows(spec))
        return all_rows, sheet_name

    rows = build_rows(document)
    sheet_name = document.get("page", {}).get("sheet_name", "埋点提报")
    return rows, sheet_name


def summarize_document(document: dict, rows: list[list]) -> str:
    if "pages" in document:
        view_n = sum(len(p.get("view_events", [])) for p in document["pages"])
        click_n = sum(len(p.get("click_events", [])) for p in document["pages"])
        show_n = sum(len(p.get("show_events", [])) for p in document["pages"])
        return (
            f"页面数: {len(document['pages'])} | "
            f"浏览 {view_n} | 点击 {click_n} | 曝光 {show_n} | 合计 {len(rows)}"
        )

    view_n = len(document.get("view_events", []))
    click_n = len(document.get("click_events", []))
    show_n = len(document.get("show_events", []))
    page = document.get("page", {})
    return (
        f"页面: {page.get('name_cn', '')} ({page.get('name_en', '')})\n"
        f"浏览 {view_n} | 点击 {click_n} | 曝光 {show_n} | 合计 {len(rows)}"
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate wealth-app tracking spec Excel")
    parser.add_argument("--input", "-i", required=True, help="JSON spec file path")
    parser.add_argument("--output", "-o", help="Output xlsx path")
    parser.add_argument("--sheet", help="Worksheet name")
    parser.add_argument("--remove-input", action="store_true", help="Delete input JSON after success")
    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    if not input_path.exists():
        print(f"Input not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    output_path = Path(args.output).resolve() if args.output else input_path.with_suffix(".xlsx")

    document = load_spec(input_path)

    # 硬约束：事件中文名须含 浏览/点击/曝光（存量缺动作时自动补齐）
    if "pages" in document:
        for p in document["pages"]:
            ev, changes = normalize_events_dict(p.get("events") or {})
            p["events"] = ev
            for c in changes:
                print(f"[event_name_cn] {c}", file=sys.stderr)
    elif "events" in document:
        ev, changes = normalize_events_dict(document.get("events") or {})
        document["events"] = ev
        for c in changes:
            print(f"[event_name_cn] {c}", file=sys.stderr)

    errs = check_spec(document)
    if errs:
        for e in errs:
            print(e, file=sys.stderr)
        sys.exit(1)

    warn_unknown_event_prefixes(document)

    rows, default_sheet = collect_rows_from_document(document)
    if not rows:
        print("No tracking events found in spec.", file=sys.stderr)
        sys.exit(1)

    sheet_name = args.sheet or default_sheet
    write_excel(rows, output_path, sheet_name=sheet_name)
    print(summarize_document(document, rows))
    print(f"Saved: {output_path}")

    if args.remove_input:
        input_path.unlink(missing_ok=True)
        print(f"Removed intermediate JSON: {input_path}")


if __name__ == "__main__":
    main()
