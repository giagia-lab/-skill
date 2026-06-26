#!/usr/bin/env python3
"""Parse department tracking Excel into a searchable registry JSON."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing dependency: openpyxl. Install with: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

# Column headers (with or without leading *)
COL = {
    "category": ("埋点分类",),
    "event_cn": ("事件名称",),
    "event_en": ("*事件英文名称", "事件英文名称"),
    "page_cn": ("页面名称",),
    "page_en": ("*页面英文名称", "页面英文名称"),
    "from_page": ("*前导页面", "前导页面"),
    "module_cn": ("模块名称",),
    "module_en": ("*模块英文名称", "模块英文名称"),
    "element_cn": ("元素名称",),
    "element_en": ("*元素英文名称", "元素英文名称"),
    "product_code": ("*产品代码", "产品代码"),
    "remarks": ("*备注", "备注"),
    "legacy_id": ("老埋点关联 eventId",),
}


def norm_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip().replace("*", "")


def build_header_map(row) -> dict[str, int]:
    headers = {norm_header(cell.value): idx for idx, cell in enumerate(row) if cell.value}
    mapping: dict[str, int] = {}
    for key, candidates in COL.items():
        for name in candidates:
            plain = name.replace("*", "")
            if plain in headers:
                mapping[key] = headers[plain]
                break
    return mapping


def cell(row, idx: int | None):
    if idx is None or idx >= len(row):
        return ""
    value = row[idx].value
    return "" if value is None else str(value).strip()


def parse_sheet(ws, sheet_name: str) -> dict:
    rows = list(ws.iter_rows())
    if not rows:
        return {"sheet_name": sheet_name, "pages": {}, "event_families": {}, "entries": []}

    header = build_header_map(rows[0])
    pages: dict[str, dict] = {}
    event_families: dict[str, dict[str, str]] = {}
    entries: list[dict] = []

    for row in rows[1:]:
        category = cell(row, header.get("category"))
        page_en = cell(row, header.get("page_en"))
        page_cn = cell(row, header.get("page_cn"))
        if not page_en and not page_cn:
            continue

        event_en = cell(row, header.get("event_en"))
        event_cn = cell(row, header.get("event_cn"))
        entry = {
            "sheet_name": sheet_name,
            "category": category,
            "event_cn": event_cn,
            "event_en": event_en,
            "page_cn": page_cn,
            "page_en": page_en,
            "from_page": cell(row, header.get("from_page")),
            "module_cn": cell(row, header.get("module_cn")),
            "module_en": cell(row, header.get("module_en")),
            "element_cn": cell(row, header.get("element_cn")),
            "element_en": cell(row, header.get("element_en")),
            "product_code": cell(row, header.get("product_code")),
            "remarks": cell(row, header.get("remarks")),
            "legacy_id": cell(row, header.get("legacy_id")),
        }
        entries.append(entry)

        if page_en and page_en not in pages:
            pages[page_en] = {"name_cn": page_cn, "sheet_name": sheet_name}

        if event_en and category in ("浏览事件", "曝光事件", "点击事件"):
            kind = {"浏览事件": "view", "曝光事件": "show", "点击事件": "click"}[category]
            fam = event_families.setdefault(page_en or "_global", {})
            fam.setdefault(kind, {"name_cn": event_cn, "name_en": event_en})

    return {
        "sheet_name": sheet_name,
        "pages": pages,
        "event_families": event_families,
        "entries": entries,
    }


def merge_registry(sheets: list[dict]) -> dict:
    all_pages: dict[str, dict] = {}
    all_sheets: dict[str, dict] = {}
    all_entries: list[dict] = []
    index_by_page: dict[str, list[dict]] = {}
    index_by_element: dict[str, list[dict]] = {}

    for sheet in sheets:
        name = sheet["sheet_name"]
        all_sheets[name] = {
            "pages": sheet["pages"],
            "event_families": sheet["event_families"],
            "entry_count": len(sheet["entries"]),
        }
        all_pages.update(sheet["pages"])
        all_entries.extend(sheet["entries"])

    for entry in all_entries:
        if entry["page_en"]:
            index_by_page.setdefault(entry["page_en"], []).append(entry)
        if entry["element_en"]:
            index_by_element.setdefault(entry["element_en"], []).append(entry)

    return {
        "source": "department_tracking_excel",
        "sheet_names": [s["sheet_name"] for s in sheets],
        "sheets": all_sheets,
        "pages": all_pages,
        "entries": all_entries,
        "index": {
            "by_page_en": {k: len(v) for k, v in index_by_page.items()},
            "by_element_en": {k: len(v) for k, v in index_by_element.items()},
        },
        "lookup": {
            "by_page_en": index_by_page,
            "by_element_en": index_by_element,
        },
    }


def compact_registry(full: dict) -> dict:
    """Drop heavy lookup lists for committed registry files."""
    compact = {k: v for k, v in full.items() if k != "lookup"}
    compact["index"] = full["index"]
    return compact


def lookup_page(registry: dict, page_en: str) -> dict | None:
    for sheet in registry.get("sheets", {}).values():
        if page_en in sheet.get("pages", {}):
            page = sheet["pages"][page_en]
            return {
                "sheet_name": page.get("sheet_name") or sheet,
                "page_cn": page.get("name_cn"),
                "page_en": page_en,
                "event_families": sheet.get("event_families", {}).get(page_en, {}),
            }
    page = registry.get("pages", {}).get(page_en)
    if page:
        return {"page_cn": page.get("name_cn"), "page_en": page_en, "event_families": {}}
    return None


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse department tracking Excel to registry JSON")
    parser.add_argument("--input", "-i", required=True, help="Department tracking xlsx path")
    parser.add_argument("--output", "-o", help="Output json path (default: {input}-registry.json)")
    parser.add_argument("--compact", action="store_true", help="Omit full lookup lists (smaller file)")
    parser.add_argument("--lookup-page", help="Print matches for a page_en")
    parser.add_argument("--lookup-element", help="Print matches for an element_en")
    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    if not input_path.exists():
        print(f"Input not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(input_path, read_only=True, data_only=True)
    sheets = [parse_sheet(wb[name], name) for name in wb.sheetnames]
    registry = merge_registry(sheets)

    if args.lookup_page or args.lookup_element:
        if args.lookup_page:
            hits = registry["lookup"]["by_page_en"].get(args.lookup_page, [])
            print(json.dumps(hits, ensure_ascii=False, indent=2))
        if args.lookup_element:
            hits = registry["lookup"]["by_element_en"].get(args.lookup_element, [])
            print(json.dumps(hits, ensure_ascii=False, indent=2))
        return

    output = registry if not args.compact else compact_registry(registry)
    output_path = Path(args.output).resolve() if args.output else input_path.with_name(
        input_path.stem + "-registry.json"
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(
        f"Sheets: {len(sheets)} | Pages: {len(registry['pages'])} | "
        f"Entries: {len(registry['entries'])}"
    )
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()
